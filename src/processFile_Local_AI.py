import os
import time
import json
import glob
import httpx
import logging
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import gc
import subprocess
from multiprocessing import Pool, current_process

# === CONFIGURATION ===
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# --- AJUSTE FEITO AQUI ---
# Aponta especificamente para a pasta onde o seletor guardou os ficheiros
INPUT_DIR = os.path.join(SCRIPT_DIR, "..", "data", "Jobs_Sampled_20pct")

OLLAMA_URL = "http://localhost:11434/api/generate"
MODEL_NAME = "qwen3:8b"
OUTPUT_PATH = os.path.join(SCRIPT_DIR, "..", "output", "results")
LOG_PATH = os.path.join(SCRIPT_DIR, "..", "logs")

temperature = 0
COOLDOWN = 0.5
NUM_PREDICT = 200
MAX_RETRIES = 2
RETRY_SLEEP = 3
BATCH_SIZE = 20
BATCH_SAVE_DIR = os.path.join(OUTPUT_PATH, "batch_temp")
NUM_WORKERS = 4  # 4 Instâncias simultâneas

os.makedirs(OUTPUT_PATH, exist_ok=True)
os.makedirs(LOG_PATH, exist_ok=True)
os.makedirs(BATCH_SAVE_DIR, exist_ok=True)

LOG_FILE = os.path.join(LOG_PATH, f"process_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.FileHandler(LOG_FILE, encoding='utf-8')]
)


# --- FUNÇÃO PARA INICIAR OLLAMA AUTOMATICAMENTE ---
def ensure_ollama_running():
    """
    Verifica se o Ollama está a correr. Se não, tenta iniciar com paralelismo configurado.
    """
    try:
        logging.info("A verificar o estado do Ollama...")
        # Tenta uma conexão simples para ver se já está ativo
        response = httpx.get("http://localhost:11434", timeout=2)
        if response.status_code == 200:
            logging.info("Ollama já está a correr.")
            return True
    except:
        logging.info("Ollama não detetado. A tentar iniciar servidor...")

    try:
        # Configura ambiente para permitir 4 requisições paralelas
        ollama_env = os.environ.copy()
        ollama_env["OLLAMA_NUM_PARALLEL"] = str(NUM_WORKERS)
        ollama_env["OLLAMA_MAX_LOADED_MODELS"] = "1"  # Mantém 1 modelo na VRAM para economizar memória

        # Inicia o processo em background
        subprocess.Popen(["ollama", "serve"], env=ollama_env, shell=True)

        print("A iniciar Ollama (a aguardar 10s)...")
        time.sleep(10)  # Tempo para o servidor subir
        return True
    except Exception as e:
        logging.error(f"Não foi possível iniciar o Ollama automaticamente: {e}")
        print("ERRO: Não foi possível iniciar o Ollama. Certifique-se que ele está instalado.")
        return False


# --- FUNÇÕES ORIGINAIS DO SEU COLEGA (MANTIDAS) ---

def read_file_adaptive(file_path, **kwargs):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    file_extension = os.path.splitext(file_path)[1].lower()

    readers = {
        '.csv': pd.read_csv,
        '.tsv': lambda path, **args: pd.read_csv(path, sep='\t', **args),
        '.xlsx': pd.read_excel, '.xls': pd.read_excel, '.xlsm': pd.read_excel,
        '.xlsb': pd.read_excel, '.odf': pd.read_excel, '.ods': pd.read_excel,
        '.odt': pd.read_excel, '.json': pd.read_json,
        '.parquet': lambda path, **args: pd.read_parquet(path, **args),
        '.parq': lambda path, **args: pd.read_parquet(path, **args),
        '.pkl': pd.read_pickle, '.pickle': pd.read_pickle,
    }

    if file_extension not in readers:
        try:
            return _detect_and_read_by_content(file_path, **kwargs)
        except:
            raise ValueError(f"Unsupported file format: {file_extension}")

    reader_func = readers[file_extension]
    try:
        return reader_func(file_path, **kwargs)
    except Exception as e:
        logging.error(f"Error reading file {file_path}: {str(e)}")
        raise e


def _detect_and_read_by_content(file_path, **kwargs):
    with open(file_path, 'rb') as f:
        header = f.read(1024)
    header_str = header.decode('utf-8', errors='ignore').strip()
    if header_str.startswith('{') or header_str.startswith('['):
        return pd.read_json(file_path, **kwargs)
    if ',' in header_str or '\t' in header_str:
        return pd.read_csv(file_path, **kwargs)
    raise ValueError(f"File format not automatically detected: {file_path}")


def condense_description(description, window=3, min_length=2000):
    if description is None or len(description) < min_length:
        return description if description is not None else ""
    keywords = ["schedule", "flexible", "flexibility", "shift", "weekend", "weekends", "holiday", "holidays", "night",
                "nights", "irregular", "on-call", "availability", "as needed", "rotating", "rotational", "hours",
                "hourly", "mandatory", "split shift", "split-shift", "variable", "overtime", "call-in", "unpredictable",
                "as required", "required to", "vary", "subject to", "full availability", "prn"]
    lines = description.split('\n')
    lines = [l for l in lines if l.strip() != ""]
    keyword_matches = set()
    for i, line in enumerate(lines):
        if any(kw in line.lower() for kw in keywords):
            for j in range(max(0, i - window), min(len(lines), i + window + 1)):
                keyword_matches.add(j)
    if not keyword_matches:
        return description
    return "\n".join(lines[i] for i in sorted(keyword_matches))


def build_flexibility_prompt(description):
    return f"""
You are an expert HR analyst specializing in evaluating job posting flexibility requirements. Your task is to analyze the job description and classify the flexibility aspects.

RESPONSE FORMAT:
Respond ONLY in the following exact JSON format:
{{
  "undesired_flexibility": "YES" or "NO",
  "undesired_quote": "exact quote or 'N/A'",
  "desired_flexibility": "YES" or "NO",
  "desired_quote": "exact quote or 'N/A'",
  "reasoning": "Your step-by-step reasoning, maximum 200 characters"
}}

CLASSIFICATION CRITERIA:

Undesired Flexibility:
- Mark as "YES" ONLY if there is a direct phrase proving the employer can unpredictably change work hours
- Examples include: "schedule may vary", "rotating shifts", "on-call required", "as needed", "PRN", "must be available for different shifts", "subject to change", "open availability required", "weekend/holiday work required"
- Demanded flexible work schedule is considered undesired flexibility, as it is open to the employer deciding working hours, with words like "must" and "necessary" hinting at that. Examples include "must be available to work flexible schedule", "must have a flexible schedule", "flexible schedule needed", "necessary flexible schedule"
- Fixed schedules like "weekend coverage" or "night shift" are NOT undesired flexibility
- The quote must be the exact phrase that justifies the classification

Desired Flexibility:
- Mark as "YES" ONLY if the employee can clearly choose when to work
- Examples include: "flexible schedule", "choose your hours", "work when you want", "set your own schedule"
- The quote must be the exact phrase that justifies the classification

Instructions:
1. Read the job description carefully
2. Identify phrases related to work schedule flexibility
3. Classify according to the criteria above
4. Provide exact quotes to support your classifications
5. Explain your reasoning step-by-step
6. Respond ONLY in the specified JSON format

Job Description:
{description}
"""


def call_ollama_api(prompt, max_retries=MAX_RETRIES, retry_sleep=RETRY_SLEEP):
    for attempt in range(max_retries):
        try:
            response = httpx.post(OLLAMA_URL, json={"model": MODEL_NAME, "prompt": prompt, "temperature": temperature,
                                                    "stream": False, "repeat_penalty": 1.2, "top_k": 50, "top_p": 0.9},
                                  timeout=120)
            if response.status_code == 200:
                return response.json()["response"]
            else:
                logging.warning(f"Status {response.status_code}: {response.text}")
        except Exception as e:
            logging.error(f"Error calling Ollama: {e}")
        time.sleep(retry_sleep)
    return None


def safe_parse_json(llm_output):
    if not llm_output: return None
    import re
    match = re.search(r'(\{[\s\S]+\})', llm_output)
    if match:
        json_str = match.group(1)
        if json_str.strip() == "{}": return {}
        try:
            return json.loads(json_str)
        except json.JSONDecodeError:
            json_str = json_str.replace("'", '"');
            json_str = re.sub(r',(\s*[}\]])', r'\1', json_str)
            try:
                return json.loads(json_str)
            except:
                pass
    return None


def validate_response(parsed_response):
    if not parsed_response: return False
    required_keys = ["undesired_flexibility", "undesired_quote", "desired_flexibility", "desired_quote", "reasoning"]
    if not all(key in parsed_response for key in required_keys): return False
    flexibility_values = [parsed_response["undesired_flexibility"], parsed_response["desired_flexibility"]]
    if not all(val in ["YES", "NO"] for val in flexibility_values): return False
    return True


def yesno_to_dummy(val):
    if isinstance(val, str):
        val = val.strip().upper()
        if val == "YES": return 1
        if val == "NO": return 0
    return 0


def get_resume_index(save_path_prefix):
    batch_files = glob.glob(f"{save_path_prefix}_*.xlsx")
    if not batch_files: return -1

    def extract(fn):
        try:
            return int(fn.split('_')[-1].split('.')[0])
        except:
            return -1

    batch_files.sort(key=extract)
    try:
        df = pd.read_excel(batch_files[-1])
        if df.empty: return -1
        match = __import__('re').search(r"\(Row_(\d+)\)", str(df.iloc[-1].get("Title", "")))
        if match: return int(match.group(1))
    except:
        pass
    return -1


def save_batch_chunk(batch_data, batch_index, save_path_prefix):
    if not batch_data: return
    df = pd.DataFrame(batch_data)
    save_path = f"{save_path_prefix}_{batch_index}.xlsx"
    try:
        os.makedirs(os.path.dirname(save_path), exist_ok=True)
        df.to_excel(save_path, index=False)
        logging.info(f"Batch {batch_index} saved.")
    except Exception as e:
        logging.error(f"Failed save: {e}")


def load_all_processed_batches(save_path_prefix, exclude_final=False):
    batch_files = glob.glob(f"{save_path_prefix}_*.xlsx")
    if exclude_final: batch_files = [f for f in batch_files if not f.endswith("_final.xlsx")]

    def extract(fn):
        try:
            return int(fn.split('_')[-1].split('.')[0]) if not fn.endswith("_final.xlsx") else float('inf')
        except:
            return -1

    batch_files.sort(key=extract)
    all_records = [];
    seen = set()
    for bf in batch_files:
        try:
            df = pd.read_excel(bf)
            for _, r in df.iterrows():
                k = r.get("Title", "")
                if k not in seen:
                    seen.add(k);
                    all_records.append(r.to_dict())
        except:
            pass
    return pd.DataFrame(all_records) if all_records else pd.DataFrame()


def color_excel(path, col="undesired_flexibility"):
    try:
        wb = load_workbook(path);
        ws = wb.active
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        col_idx = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == col: col_idx = idx; break
        if col_idx:
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value == 1:
                        cell.fill = red
                    elif cell.value == 0:
                        cell.fill = green
        wb.save(path)
    except:
        pass


# --- PROCESSAMENTO PRINCIPAL (ADAPTADO PARA MEMÓRIA E WORKERS) ---
def process_job_postings(input_path):
    worker_name = current_process().name
    filename = os.path.basename(input_path)

    # Remove extensões para o nome base
    filename_clean = filename.replace('.csv.gz', '').replace('.csv', '')

    # Define caminho de output
    batch_save_prefix = os.path.join(BATCH_SAVE_DIR, f"{filename_clean}_batch")
    final_file_path = os.path.join(OUTPUT_PATH, f"{filename_clean}_processed.xlsx")

    # Skip se já processado
    if os.path.exists(final_file_path):
        print(f"[{worker_name}] Arquivo já processado: {filename}. A ignorar.")
        return True

    logging.info(f"[{worker_name}] A iniciar: {input_path}")

    # Resume
    last_idx = get_resume_index(batch_save_prefix)
    resume_from = last_idx + 1
    if resume_from > 0:
        logging.info(f"[{worker_name}] A retomar da linha {resume_from}")

    # Leitura com Chunking (Mantendo compatibilidade com código do colega)
    is_csv = input_path.lower().endswith(('.csv', '.csv.gz', '.tsv'))
    try:
        if is_csv:
            iterator = pd.read_csv(input_path, chunksize=500)
        else:
            df_full = read_file_adaptive(input_path)
            # Simula chunking com 1 bloco único
            df_full.columns = [c.upper() for c in df_full.columns]
            iterator = [df_full]
    except Exception as e:
        logging.error(f"Erro ao ler {input_path}: {e}")
        return False

    current_row_global_idx = 0
    current_batch = []
    current_batch_index = (resume_from // BATCH_SIZE) + 1

    for chunk_df in iterator:
        if is_csv: chunk_df.columns = [c.upper() for c in chunk_df.columns]

        # Otimização: se o chunk inteiro já foi processado, pule
        if current_row_global_idx + len(chunk_df) <= resume_from:
            current_row_global_idx += len(chunk_df)
            del chunk_df;
            gc.collect()
            continue

        for _, row in chunk_df.iterrows():
            if current_row_global_idx < resume_from:
                current_row_global_idx += 1;
                continue

            # --- LÓGICA DO SEU COLEGA ---
            desc = row.get("BODY")
            short_desc = condense_description(desc)
            prompt = build_flexibility_prompt(short_desc)

            validated = False;
            attempts = 0
            response = None;
            parsed = None

            while not validated and attempts < 3:
                attempts += 1
                response = call_ollama_api(prompt)
                parsed = safe_parse_json(response)
                validated = validate_response(parsed)

            if not validated: response = str(response) + " [FAILED VALIDATION]"

            # Monta registro
            record = {
                "Title": f"{row.get('TITLE_NAME', '')} (Row_{current_row_global_idx})",
                "Body": desc,
                "llama_raw_response": response,
                "undesired_flexibility": yesno_to_dummy(parsed.get("undesired_flexibility")) if parsed else 0,
                "desired_flexibility": yesno_to_dummy(parsed.get("desired_flexibility")) if parsed else 0,
                "undesired_quote": parsed.get("undesired_quote") if parsed else "",
                "desired_quote": parsed.get("desired_quote") if parsed else "",
                "reasoning": parsed.get("reasoning") if parsed else "PARSING ERROR"
            }

            current_batch.append(record)
            current_row_global_idx += 1

            # GC do colega (dentro do loop)
            if len(current_batch) >= BATCH_SIZE:
                save_batch_chunk(current_batch, current_batch_index, batch_save_prefix)
                current_batch = [];
                current_batch_index += 1;
                gc.collect()

        # GC nosso (limpa chunk da memória)
        del chunk_df;
        gc.collect()

    # Salva resto
    if current_batch:
        save_batch_chunk(current_batch, current_batch_index, batch_save_prefix)
        del current_batch;
        gc.collect()

    # Finaliza
    df_final = load_all_processed_batches(batch_save_prefix, exclude_final=True)
    if not df_final.empty:
        if "Title" in df_final.columns:
            df_final["Title"] = df_final["Title"].str.replace(r" \(Row_\d+\)", "", regex=True)
        df_final.to_excel(final_file_path, index=False)
        color_excel(final_file_path)

    print(f"[{worker_name}] Finalizado: {filename}")
    return True


# Wrapper para o Pool não quebrar com exceções
def process_wrapper(f):
    try:
        return process_job_postings(f)
    except Exception as e:
        print(f"Erro crítico no arquivo {f}: {e}")
        return False


def ollama_warmup():
    try:
        call_ollama_api("Test", max_retries=1)
    except:
        pass


def run_batch_processing():
    # Busca recursiva na pasta INPUT_DIR que agora aponta para Jobs_Sampled_20pct
    search_pattern = os.path.join(INPUT_DIR, "**", "*.csv.gz")
    files = glob.glob(search_pattern, recursive=True)

    if not files:  # Tenta sem .gz
        files = glob.glob(os.path.join(INPUT_DIR, "**", "*.csv"), recursive=True)

    if not files:
        print(f"Nenhum arquivo encontrado em {INPUT_DIR}")
        return

    print(f"Encontrados {len(files)} arquivos. A iniciar {NUM_WORKERS} workers.")

    # Processamento paralelo
    with Pool(processes=NUM_WORKERS) as pool:
        pool.map(process_wrapper, files)

    print("Processamento total finalizado.")


if __name__ == "__main__":
    # 1. Tenta garantir que o Ollama está a correr com suporte a paralelismo
    ensure_ollama_running()

    # 2. Warmup
    ollama_warmup()

    # 3. Executa
    run_batch_processing()

    logging.info(f"Log salvo em: {LOG_FILE}")